"""
팀원 비밀평가(동료평가) — 1회성 일괄 진행
- 관리자: /비밀평가 시작 → 과제-대시보드 채널에 패널 게시(팀별 완료율 + [✍️ 평가하기])
          /비밀평가 결과 → 대상자별 평균 점수·코멘트 집계(관리자 전용, 엑셀 포함)
          /비밀평가 종료 → 라운드 마감
- 수강생: [✍️ 평가하기] → 같은 팀원 선택(본인 제외) → Modal에 4지표 1~5점 + 코멘트 → 저장
- 비밀평가: 평가자 정보는 저장은 하되(중복방지·진행률) 집계 결과에는 절대 노출하지 않는다.
- 팀 판별: 닉네임 괄호/언더바 안의 팀 번호로 자동 인식 — 예) 송승현(3팀), 남영찬_팀3, 팀3
- 대시보드에는 팀별 '완료율'만 공개하며 점수는 관리자만 확인한다.
"""
import io
import logging
import re

import discord
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from discord import app_commands
from discord.ext import commands

import config
import database
from ui.embeds import KST

log = logging.getLogger("asanAX.peer_eval")

PANEL_TYPE = "peer_eval"
TEAMS = list(config.TEAM_CHANNELS.keys())  # 단일 출처: config.TEAM_CHANNELS
ACCENT = "#8B5CF6"  # 비밀평가 전용 보라 계열

# ── 채점 지표(참조 루브릭 4개 × 5점 척도) ──────────────────────────────────────
# (마커, 지표명, 관찰 근거, (5점 우수, 3점 보통, 1점 미흡))
INDICATORS: list[tuple[str, str, str, tuple[str, str, str]]] = [
    (
        "①", "과업 완수(산출물)",
        "맡은 역할의 결과물을 실제로 냈는가 (PR·커밋, 워크시트·PRD, 데모 자료, 인터뷰 등)",
        (
            "맡은 산출물을 기한 내 완성했고, 품질이 팀 데모·제출에 직접 기여함",
            "대부분 완수했으나 일부는 지연되거나 다른 팀원이 보완함",
            "산출물이 거의 없거나, 맡은 일을 다른 팀원이 대신함",
        ),
    ),
    (
        "②", "협업·소통",
        "회의·토론에 적극 참여하고 진행상황을 투명하게 공유, 연락·리뷰에 응답을 잘함",
        (
            "세션·토론에 능동 참여하고 진행상황·막힘을 먼저 공유함",
            "요청하면 응답하지만 먼저 공유하는 일은 적음",
            "회의 불참·잠수, 연락이 잘 닿지 않아 팀 진행이 막힘",
        ),
    ),
    (
        "③", "책임·약속 이행",
        "마감·팀 결정사항 준수, 맡은 일을 끝까지 마무리하는 신뢰성",
        (
            "기한·약속을 항상 지켜 리마인드가 필요 없음",
            "가끔 지연되지만 사전에 알리고 수습함",
            "상습 지연·무단 미이행으로 팀 일정에 피해를 줌",
        ),
    ),
    (
        "④", "팀 기여·문제해결",
        "블로커 해결, PR 리뷰·지식 공유, 더 나은 방안 제안 등 팀 전체를 끌어올린 정도",
        (
            "문제 해결·리뷰·지식 공유로 팀 성과를 견인함",
            "자기 몫은 하지만 팀 차원의 기여는 제한적",
            "팀 문제에 무관심하고 자기 영역 밖 기여가 없음",
        ),
    ),
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_admin(interaction: discord.Interaction) -> bool:
    member = interaction.user
    if not isinstance(member, discord.Member):
        return False
    return any(r.id == config.ADMIN_ROLE_ID for r in member.roles)


def _is_staff(member: discord.Member) -> bool:
    return any(r.id == config.ADMIN_ROLE_ID for r in member.roles)


def parse_team(display_name: str) -> str | None:
    """닉네임에서 팀 번호를 추출해 config.TEAM_CHANNELS 키(팀N)로 정규화.
    지원 형식: 홍길동(3팀), 홍길동(팀3), 남영찬_팀3, 팀3, 3조 등."""
    m = re.search(r"(\d+)\s*[팀조]", display_name) or re.search(r"[팀조]\s*(\d+)", display_name)
    if not m:
        return None
    key = f"팀{m.group(1)}"
    return key if key in config.TEAM_CHANNELS else None


def build_roster(guild: discord.Guild) -> dict[str, list[discord.Member]]:
    """닉네임 기준 팀별 수강생 명단. 봇·관리자(스태프)는 평가 대상에서 제외."""
    roster: dict[str, list[discord.Member]] = {t: [] for t in TEAMS}
    for m in guild.members:
        if m.bot or _is_staff(m):
            continue
        team = parse_team(m.display_name)
        if team:
            roster[team].append(m)
    for team in roster:
        roster[team].sort(key=lambda mm: mm.display_name)
    return roster


def _bar(done: int, total: int, width: int = 10) -> str:
    if total <= 0:
        return "─" * width
    filled = round(width * done / total)
    return "█" * filled + "░" * (width - filled)


# ── Progress (완료율) ──────────────────────────────────────────────────────────

def compute_progress(
    roster: dict[str, list[discord.Member]], evals: list[dict]
) -> dict[str, tuple[int, int]]:
    """팀별 (완료 인원, 전체 인원). 팀원 전원(본인 제외)을 평가한 사람만 '완료'."""
    # evaluator_id -> {target_id...}
    done_map: dict[str, set[str]] = {}
    for e in evals:
        done_map.setdefault(e["evaluator_id"], set()).add(e["target_id"])

    result: dict[str, tuple[int, int]] = {}
    for team, members in roster.items():
        member_ids = {str(m.id) for m in members}
        total = len(members)
        if total <= 1:
            # 평가할 팀원이 없으면 전원 완료로 간주(1명 이하 팀)
            result[team] = (total, total)
            continue
        completed = 0
        for m in members:
            expected = member_ids - {str(m.id)}
            done = done_map.get(str(m.id), set()) & expected
            if len(done) >= len(expected):
                completed += 1
        result[team] = (completed, total)
    return result


def build_panel_embed(
    round_row: dict,
    roster: dict[str, list[discord.Member]],
    evals: list[dict],
) -> discord.Embed:
    active = bool(round_row["is_active"])
    now_str = discord.utils.utcnow().astimezone(KST).strftime("%Y-%m-%d %H:%M KST")

    status = "진행 중" if active else "마감됨"
    embed = discord.Embed(
        title=f"🔒 팀원 비밀평가 — {round_row['title']}",
        description=(
            f"상태: **{status}** · 각 조 팀원끼리 서로를 평가합니다.\n"
            "**평가 내용은 비밀입니다.** 누가 몇 점을 줬는지는 공개되지 않으며, "
            "집계 결과는 관리자만 확인합니다.\n\n"
            + ("아래 **[✍️ 평가하기]** 버튼으로 같은 팀원을 평가하세요."
               if active else "이 라운드는 마감되어 더 이상 평가할 수 없습니다.")
        ),
        color=discord.Color.from_str(ACCENT),
    )

    # 채점 지표 요약
    rubric = "\n".join(f"{mark} **{name}** — {basis}" for mark, name, basis, _ in INDICATORS)
    embed.add_field(
        name="📋 채점 지표 (각 1~5점 · 5·3·1 앵커, 4·2점은 중간)",
        value=rubric,
        inline=False,
    )

    # 팀별 완료율 (점수는 비공개 — 완료 인원만)
    progress = compute_progress(roster, evals)
    lines = []
    for team in TEAMS:
        done, total = progress.get(team, (0, 0))
        if total == 0:
            lines.append(f"`{_bar(0, 1)}` **{team}** — 명단 없음")
        else:
            pct = round(100 * done / total)
            lines.append(f"`{_bar(done, total)}` **{team}** — {done}/{total}명 완료 ({pct}%)")
    embed.add_field(name="✅ 팀별 진행 현황", value="\n".join(lines), inline=False)

    embed.set_footer(text=f"아산 AX · 비밀평가 · {now_str}")
    return embed


async def refresh_panel(bot: commands.Bot) -> bool:
    """저장된 패널 메시지를 최신 진행 현황으로 edit."""
    panel = await database.get_assignment_panel(PANEL_TYPE)
    if not panel:
        return False
    guild = bot.get_guild(config.GUILD_ID)
    if not guild:
        return False
    ch = guild.get_channel(int(panel["channel_id"]))
    if not ch or not isinstance(ch, discord.TextChannel):
        return False

    round_row = await database.get_active_peer_round()
    if not round_row:
        # 활성 라운드가 없으면 마지막 라운드로 표시만 갱신
        return False

    try:
        msg = await ch.fetch_message(int(panel["message_id"]))
        evals = await database.get_peer_evaluations(round_row["id"])
        roster = build_roster(guild)
        embed = build_panel_embed(round_row, roster, evals)
        await msg.edit(embed=embed, view=PeerEvalPanelView(bot))
        return True
    except (discord.NotFound, discord.HTTPException) as e:
        log.warning("Peer-eval panel refresh failed: %s", e)
        return False


# ── 게시 헬퍼 (진행판 · 팀 채널 패널) ───────────────────────────────────────────

def build_team_panel_embed(round_row: dict, team: str) -> discord.Embed:
    """각 팀 채널에 게시하는 평가 진입 패널(팀별 안내 + 채점 기준)."""
    embed = discord.Embed(
        title=f"🔒 {team} 팀원 비밀평가 — {round_row['title']}",
        description=(
            "같은 팀원끼리 서로를 평가합니다. 아래 **[✍️ 평가하기]** 버튼을 눌러 "
            "팀원별로 4개 지표를 각각 1~5점으로 매겨주세요.\n"
            "**평가 내용은 비밀입니다.** 누가 몇 점을 줬는지는 공개되지 않으며, "
            "집계 결과는 관리자만 확인합니다. 제출 후에도 언제든 수정할 수 있습니다."
        ),
        color=discord.Color.from_str(ACCENT),
    )
    rubric = "\n".join(
        f"{mark} **{name}** · 5점: {anchors[0]} / 1점: {anchors[2]}"
        for mark, name, _basis, anchors in INDICATORS
    )
    embed.add_field(
        name="📋 채점 지표 (각 1~5점 · 5·3·1 앵커, 4·2점은 중간)",
        value=rubric[:1024],
        inline=False,
    )
    embed.set_footer(text="아산 AX · 비밀평가")
    return embed


async def post_progress_dashboard(
    bot: commands.Bot, guild: discord.Guild, round_row: dict
) -> discord.TextChannel | None:
    """과제-대시보드 채널에 진행판을 게시(또는 갱신)하고 추적 정보를 저장. 채널 반환."""
    ch = guild.get_channel(config.ASSIGNMENT_DASHBOARD_CHANNEL_ID)
    if not ch or not isinstance(ch, discord.TextChannel):
        return None

    roster = build_roster(guild)
    evals = await database.get_peer_evaluations(round_row["id"])
    embed = build_panel_embed(round_row, roster, evals)

    # 이미 추적 중인 진행판이 있으면 갱신, 없으면 새로 게시
    panel = await database.get_assignment_panel(PANEL_TYPE)
    if panel and int(panel["channel_id"]) == ch.id:
        try:
            msg = await ch.fetch_message(int(panel["message_id"]))
            await msg.edit(embed=embed, view=PeerEvalPanelView(bot))
            return ch
        except (discord.NotFound, discord.HTTPException):
            pass

    msg = await ch.send(embed=embed, view=PeerEvalPanelView(bot))
    await database.save_assignment_panel(PANEL_TYPE, str(ch.id), str(msg.id))
    return ch


async def post_team_channel_panels(
    bot: commands.Bot, guild: discord.Guild, round_row: dict
) -> list[str]:
    """각 팀 채널에 평가 진입 패널을 게시. 성공한 팀명 목록 반환."""
    posted: list[str] = []
    for team, channel_id in config.TEAM_CHANNELS.items():
        ch = guild.get_channel(channel_id)
        if not ch or not isinstance(ch, discord.TextChannel):
            log.warning("Team channel not found for peer-eval: %s (%d)", team, channel_id)
            continue
        try:
            await ch.send(
                embed=build_team_panel_embed(round_row, team),
                view=PeerEvalPanelView(bot),
            )
            posted.append(team)
        except discord.HTTPException as e:
            log.warning("Failed to post peer-eval panel to %s: %s", team, e)
    return posted


async def get_or_create_active_round(default_title: str = "팀원 상호평가") -> dict:
    round_row = await database.get_active_peer_round()
    if round_row:
        return round_row
    rid = await database.create_peer_round(default_title)
    created = await database.get_peer_round(rid)
    assert created is not None
    return created


# ── 지표별 스텝 평가 흐름 ───────────────────────────────────────────────────────

class CommentModal(discord.ui.Modal, title="코멘트 입력"):
    """검토 단계에서 코멘트만 따로 입력받는 모달."""

    def __init__(self, flow: "EvalFlowView") -> None:
        super().__init__()
        self.flow = flow
        self._input = discord.ui.TextInput(
            label="코멘트 (선택 · 비밀 · 관리자만 열람)",
            style=discord.TextStyle.paragraph,
            placeholder="구체적 근거나 피드백을 자유롭게 적어주세요.",
            default=flow.comment,
            required=False,
            max_length=500,
        )
        self.add_item(self._input)

    async def on_submit(self, interaction: discord.Interaction) -> None:
        self.flow.comment = self._input.value.strip()
        self.flow.build()
        await interaction.response.edit_message(embed=self.flow.embed(), view=self.flow)


class EvalFlowView(discord.ui.View):
    """대상자 1명을 지표별 스텝(1/4~4/4)으로 평가 → 검토 → 저장.
    각 스텝에서 해당 지표의 전체 채점 기준(5·3·1 앵커)을 임베드로 보여준다."""

    def __init__(
        self,
        bot: commands.Bot,
        round_id: int,
        team: str,
        target: discord.Member,
        existing: dict | None = None,
        test: bool = False,
    ) -> None:
        super().__init__(timeout=600)
        self.bot = bot
        self.round_id = round_id
        self.team = team
        self.target = target
        self.test = test
        self.scores: list[int | None] = (
            [existing[f"score{i + 1}"] for i in range(4)] if existing else [None, None, None, None]
        )
        self.comment: str = existing["comment"] if existing else ""
        self.index = 0
        self.build()

    # ── 렌더링 ──────────────────────────────────────────────────────────────
    def build(self) -> None:
        self.clear_items()
        if self.index < 4:
            _mark, name, _basis, anchors = INDICATORS[self.index]
            labels = [
                ("5", "5점 · 우수", anchors[0]),
                ("4", "4점 · 우수와 보통 사이", "중간 수준"),
                ("3", "3점 · 보통", anchors[1]),
                ("2", "2점 · 보통과 미흡 사이", "중간 수준"),
                ("1", "1점 · 미흡", anchors[2]),
            ]
            opts = []
            for value, label, desc in labels:
                opt = discord.SelectOption(label=label, value=value, description=desc[:100])
                if self.scores[self.index] == int(value):
                    opt.default = True
                opts.append(opt)
            sel = discord.ui.Select(
                placeholder=f"{name} 점수를 선택하세요",
                min_values=1, max_values=1, options=opts, row=0,
            )
            sel.callback = self._on_score
            self.add_item(sel)
            if self.index > 0:
                back = discord.ui.Button(label="◀ 이전", style=discord.ButtonStyle.secondary, row=1)
                back.callback = self._on_back
                self.add_item(back)
        else:
            comment_btn = discord.ui.Button(
                label="💬 코멘트 " + ("수정" if self.comment else "입력"),
                style=discord.ButtonStyle.secondary, row=0,
            )
            comment_btn.callback = self._on_comment
            self.add_item(comment_btn)
            save = discord.ui.Button(label="✅ 저장", style=discord.ButtonStyle.success, row=0)
            save.callback = self._on_save
            self.add_item(save)
            back = discord.ui.Button(label="◀ 이전", style=discord.ButtonStyle.secondary, row=1)
            back.callback = self._on_back
            self.add_item(back)

    def embed(self) -> discord.Embed:
        prefix = "🧪 [테스트] " if self.test else ""
        chosen = " ".join(
            f"{INDICATORS[i][0]}{self.scores[i] if self.scores[i] is not None else '·'}"
            for i in range(4)
        )
        if self.index < 4:
            mark, name, basis, anchors = INDICATORS[self.index]
            e = discord.Embed(
                title=f"{prefix}[{self.index + 1}/4] {mark} {name}",
                description=f"**무엇을 보는가**\n{basis}",
                color=discord.Color.from_str(ACCENT),
            )
            e.add_field(
                name="채점 기준 (5점 척도 · 4·2점은 중간)",
                value=(
                    f"🟢 **5점 · 우수** — {anchors[0]}\n"
                    f"🟡 **3점 · 보통** — {anchors[1]}\n"
                    f"🔴 **1점 · 미흡** — {anchors[2]}"
                )[:1024],
                inline=False,
            )
            e.set_footer(text=f"대상: {self.target.display_name} · 진행 {chosen}")
        else:
            avg = sum(s for s in self.scores if s is not None) / 4
            lines = [f"{INDICATORS[i][0]} **{INDICATORS[i][1]}** — {self.scores[i]}점" for i in range(4)]
            e = discord.Embed(
                title=f"{prefix}평가 검토 — {self.target.display_name}",
                description="점수를 확인하고 **저장**하세요. 코멘트는 선택입니다. `◀ 이전`으로 수정할 수 있습니다.",
                color=discord.Color.from_str(ACCENT),
            )
            e.add_field(name=f"입력한 점수 (평균 {avg:.1f})", value="\n".join(lines), inline=False)
            if self.comment:
                e.add_field(name="코멘트", value=self.comment[:1024], inline=False)
        return e

    async def _rerender(self, interaction: discord.Interaction) -> None:
        self.build()
        await interaction.response.edit_message(embed=self.embed(), view=self)

    # ── 콜백 ────────────────────────────────────────────────────────────────
    async def _on_score(self, interaction: discord.Interaction) -> None:
        self.scores[self.index] = int(interaction.data["values"][0])  # type: ignore[index]
        self.index += 1
        await self._rerender(interaction)

    async def _on_back(self, interaction: discord.Interaction) -> None:
        self.index = max(0, self.index - 1)
        await self._rerender(interaction)

    async def _on_comment(self, interaction: discord.Interaction) -> None:
        await interaction.response.send_modal(CommentModal(self))

    async def _on_save(self, interaction: discord.Interaction) -> None:
        if any(s is None for s in self.scores):
            await interaction.response.send_message("모든 지표를 선택한 뒤 저장해주세요.", ephemeral=True)
            return
        scores = (self.scores[0], self.scores[1], self.scores[2], self.scores[3])

        # 테스트(미리보기) — 저장하지 않음
        if self.test:
            avg = sum(scores) / 4  # type: ignore[arg-type]
            await interaction.response.edit_message(
                embed=discord.Embed(
                    title="🧪 테스트 완료 — 저장되지 않았습니다",
                    description=(
                        f"대상: **{self.target.display_name}** · 평균 **{avg:.1f}**\n"
                        "실제 평가 흐름과 동일하며, 이 입력은 집계에 반영되지 않습니다."
                    ),
                    color=discord.Color.from_str(ACCENT),
                ),
                view=None,
            )
            return

        round_row = await database.get_peer_round(self.round_id)
        if not round_row or not round_row["is_active"]:
            await interaction.response.edit_message(
                content="이 비밀평가는 마감되었습니다.", embed=None, view=None
            )
            return

        await database.save_peer_evaluation(
            round_id=self.round_id,
            team=self.team,
            evaluator_id=str(interaction.user.id),
            target_id=str(self.target.id),
            target_name=self.target.display_name,
            scores=scores,  # type: ignore[arg-type]
            comment=self.comment,
        )
        await refresh_panel(self.bot)

        # 남은 팀원 목록을 이어서 제시
        guild = interaction.guild
        assert guild is not None
        teammates = [m for m in build_roster(guild).get(self.team, []) if m.id != interaction.user.id]
        done = await database.get_evaluated_target_ids(self.round_id, str(interaction.user.id))
        remaining = [m for m in teammates if str(m.id) not in done]

        avg = sum(scores) / 4  # type: ignore[arg-type]
        summary = " · ".join(f"{INDICATORS[i][0]}{scores[i]}" for i in range(4))
        desc = f"**{self.target.display_name}** 평가 저장됨 — {summary}  (평균 {avg:.1f})\n\n"
        if remaining:
            desc += f"남은 팀원: **{len(remaining)}명** — 아래에서 이어서 평가하세요."
            next_view = TargetSelectView(self.bot, self.round_id, self.team, teammates, done)
        else:
            desc += "🎉 같은 팀 전원 평가를 완료했습니다. 감사합니다!"
            next_view = None

        await interaction.response.edit_message(
            embed=discord.Embed(
                title="✅ 저장 완료",
                description=desc,
                color=discord.Color.from_str(ACCENT),
            ),
            view=next_view,
        )


# ── 팀원 선택 View ─────────────────────────────────────────────────────────────

class TargetSelectView(discord.ui.View):
    def __init__(
        self,
        bot: commands.Bot,
        round_id: int,
        team: str,
        teammates: list[discord.Member],
        done: set[str],
        test: bool = False,
    ) -> None:
        super().__init__(timeout=300)
        self.bot = bot
        self.round_id = round_id
        self.team = team
        self.test = test
        self._members = {str(m.id): m for m in teammates}

        options = []
        for m in teammates[:25]:
            checked = str(m.id) in done
            options.append(
                discord.SelectOption(
                    label=m.display_name[:100],
                    value=str(m.id),
                    description="평가 완료 — 다시 선택 시 수정" if checked else "미평가",
                    emoji="✅" if checked else "⬜",
                )
            )
        select = discord.ui.Select(
            placeholder="테스트할 팀원을 선택하세요" if test else "평가할 팀원을 선택하세요",
            min_values=1,
            max_values=1,
            options=options,
        )
        select.callback = self._on_select
        self.add_item(select)

    async def _on_select(self, interaction: discord.Interaction) -> None:
        target_id = interaction.data["values"][0]  # type: ignore[index]
        target = self._members.get(target_id)
        if target is None:
            await interaction.response.send_message(
                "대상 팀원을 찾을 수 없습니다. 다시 시도해주세요.", ephemeral=True
            )
            return

        if self.test:
            flow = EvalFlowView(self.bot, self.round_id, self.team, target, None, test=True)
            await interaction.response.send_message(embed=flow.embed(), view=flow, ephemeral=True)
            return

        # 라운드가 아직 열려 있는지 확인
        round_row = await database.get_peer_round(self.round_id)
        if not round_row or not round_row["is_active"]:
            await interaction.response.send_message(
                "이 비밀평가는 마감되었습니다.", ephemeral=True
            )
            return

        # 기존 평가가 있으면 기본값으로 채워 수정 가능하게
        evals = await database.get_peer_evaluations(self.round_id)
        existing = next(
            (
                e for e in evals
                if e["evaluator_id"] == str(interaction.user.id) and e["target_id"] == target_id
            ),
            None,
        )
        flow = EvalFlowView(self.bot, self.round_id, self.team, target, existing)
        await interaction.response.send_message(embed=flow.embed(), view=flow, ephemeral=True)


# ── 패널 View (persistent) ─────────────────────────────────────────────────────

class PeerEvalPanelView(discord.ui.View):
    """봇 재시작 후에도 동작하도록 custom_id 고정."""

    def __init__(self, bot: commands.Bot) -> None:
        super().__init__(timeout=None)
        self.bot = bot

    @discord.ui.button(
        label="✍️ 평가하기",
        style=discord.ButtonStyle.primary,
        custom_id="peereval:evaluate",
    )
    async def evaluate(
        self, interaction: discord.Interaction, button: discord.ui.Button
    ) -> None:
        round_row = await database.get_active_peer_round()
        if not round_row:
            await interaction.response.send_message(
                "현재 진행 중인 비밀평가가 없습니다.", ephemeral=True
            )
            return

        member = interaction.user
        guild = interaction.guild
        if not isinstance(member, discord.Member) or guild is None:
            await interaction.response.send_message(
                "서버에서만 사용할 수 있습니다.", ephemeral=True
            )
            return

        if _is_staff(member):
            await interaction.response.send_message(
                "관리자·멘토는 평가 대상에서 제외되어 평가에 참여하지 않습니다.", ephemeral=True
            )
            return

        team = parse_team(member.display_name)
        if not team:
            await interaction.response.send_message(
                embed=discord.Embed(
                    title="팀 정보를 찾을 수 없습니다",
                    description=(
                        "닉네임에서 소속 팀을 인식하지 못했습니다.\n"
                        "닉네임을 **`홍길동(3팀)`** 형식으로 설정한 뒤 다시 시도하거나, 관리자에게 문의하세요."
                    ),
                    color=discord.Color.orange(),
                ),
                ephemeral=True,
            )
            return

        teammates = [m for m in build_roster(guild).get(team, []) if m.id != member.id]
        if not teammates:
            await interaction.response.send_message(
                f"**{team}**에 평가할 다른 팀원이 없습니다.", ephemeral=True
            )
            return

        done = await database.get_evaluated_target_ids(round_row["id"], str(member.id))
        rubric = "\n".join(
            f"{mark} **{name}** · 5점: {anchors[0]} / 3점: {anchors[1]} / 1점: {anchors[2]}"
            for mark, name, _basis, anchors in INDICATORS
        )
        embed = discord.Embed(
            title=f"🔒 {team} 팀원 비밀평가",
            description=(
                f"같은 팀 **{len(teammates)}명**을 평가합니다. "
                f"(완료 {len(done & {str(m.id) for m in teammates})}/{len(teammates)}명)\n"
                "아래에서 팀원을 선택해 4개 지표를 각각 1~5점으로 매겨주세요.\n"
                "**평가 내용은 비밀이며, 언제든 다시 선택해 수정할 수 있습니다.**"
            ),
            color=discord.Color.from_str(ACCENT),
        )
        embed.add_field(name="채점 기준", value=rubric[:1024], inline=False)

        await interaction.response.send_message(
            embed=embed,
            view=TargetSelectView(self.bot, round_row["id"], team, teammates, done),
            ephemeral=True,
        )

    @discord.ui.button(
        label="📊 결과 보기 (관리자)",
        style=discord.ButtonStyle.secondary,
        custom_id="peereval:results",
    )
    async def results(
        self, interaction: discord.Interaction, button: discord.ui.Button
    ) -> None:
        if not _is_admin(interaction):
            await interaction.response.send_message(
                "집계 결과는 관리자만 확인할 수 있습니다.", ephemeral=True
            )
            return
        round_row = await database.get_active_peer_round()
        if not round_row:
            await interaction.response.send_message(
                "진행 중인 비밀평가가 없습니다. `/비밀평가 결과`로 조회하세요.", ephemeral=True
            )
            return
        await _send_results(interaction, self.bot, round_row)


# ── 관리자 제어판 (과제 대시보드 버튼에서 호출) ──────────────────────────────────

class TeamPickView(discord.ui.View):
    """테스트(미리보기)용 — 평가해 볼 팀을 선택."""

    def __init__(self, bot: commands.Bot, round_id: int) -> None:
        super().__init__(timeout=180)
        self.bot = bot
        self.round_id = round_id

        options = [discord.SelectOption(label=t, value=t, emoji="👥") for t in TEAMS]
        select = discord.ui.Select(
            placeholder="평가 흐름을 테스트할 팀을 선택하세요",
            min_values=1,
            max_values=1,
            options=options,
        )
        select.callback = self._on_select
        self.add_item(select)

    async def _on_select(self, interaction: discord.Interaction) -> None:
        team = interaction.data["values"][0]  # type: ignore[index]
        guild = interaction.guild
        assert guild is not None
        members = build_roster(guild).get(team, [])
        if not members:
            await interaction.response.send_message(
                f"**{team}**에 닉네임으로 인식된 팀원이 없어 테스트할 대상이 없습니다.\n"
                "팀원 닉네임을 `홍길동(3팀)` 형식으로 설정한 뒤 다시 시도하세요.",
                ephemeral=True,
            )
            return
        await interaction.response.send_message(
            embed=discord.Embed(
                title=f"🧪 {team} 평가 미리보기",
                description=(
                    "실제 수강생이 보는 화면과 동일합니다. 팀원을 선택해 점수를 입력해보세요.\n"
                    "**입력값은 저장되지 않습니다.**"
                ),
                color=discord.Color.from_str(ACCENT),
            ),
            view=TargetSelectView(self.bot, self.round_id, team, members, set(), test=True),
            ephemeral=True,
        )


class PeerEvalAdminView(discord.ui.View):
    """과제 대시보드의 [🔒 비밀평가 게시] 버튼이 여는 관리자 제어판(ephemeral)."""

    def __init__(self, bot: commands.Bot) -> None:
        super().__init__(timeout=300)
        self.bot = bot

    @discord.ui.button(label="📢 팀 채널에 게시", style=discord.ButtonStyle.success, row=0)
    async def publish_teams(
        self, interaction: discord.Interaction, button: discord.ui.Button
    ) -> None:
        guild = interaction.guild
        if not guild:
            await interaction.response.send_message("서버에서만 사용할 수 있습니다.", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True)

        round_row = await get_or_create_active_round()
        posted = await post_team_channel_panels(self.bot, guild, round_row)
        dash = await post_progress_dashboard(self.bot, guild, round_row)

        roster = build_roster(guild)
        total = sum(len(v) for v in roster.values())
        posted_line = ", ".join(posted) if posted else "게시된 채널 없음"
        dash_line = dash.mention if dash else "과제 대시보드 채널을 찾지 못함"
        await interaction.followup.send(
            embed=discord.Embed(
                title="✅ 비밀평가 게시 완료",
                description=(
                    f"라운드: **{round_row['title']}**\n"
                    f"게시된 팀 채널: **{posted_line}**\n"
                    f"진행판(완료율): {dash_line}\n\n"
                    f"인식된 명단: 총 **{total}명**. 닉네임에 팀 정보(예: `홍길동(3팀)`)가 없는 인원은 제외됩니다.\n"
                    "결과는 이 제어판의 **[📊 결과 보기]** 또는 `/비밀평가 결과`로 관리자만 확인합니다."
                ),
                color=discord.Color.green(),
            ),
            ephemeral=True,
        )

    @discord.ui.button(label="🧪 평가 미리보기(테스트)", style=discord.ButtonStyle.secondary, row=0)
    async def test_preview(
        self, interaction: discord.Interaction, button: discord.ui.Button
    ) -> None:
        round_row = await get_or_create_active_round()
        await interaction.response.send_message(
            embed=discord.Embed(
                title="🧪 평가 미리보기",
                description=(
                    "저장 없이 평가 화면을 그대로 체험할 수 있습니다.\n"
                    "먼저 테스트할 팀을 선택하세요."
                ),
                color=discord.Color.from_str(ACCENT),
            ),
            view=TeamPickView(self.bot, round_row["id"]),
            ephemeral=True,
        )

    @discord.ui.button(label="📊 결과 보기", style=discord.ButtonStyle.primary, row=1)
    async def results(
        self, interaction: discord.Interaction, button: discord.ui.Button
    ) -> None:
        round_row = await database.get_active_peer_round()
        if not round_row:
            await interaction.response.send_message(
                "진행 중인 비밀평가가 없습니다. 먼저 게시하세요.", ephemeral=True
            )
            return
        await _send_results(interaction, self.bot, round_row)

    @discord.ui.button(label="🔒 라운드 종료", style=discord.ButtonStyle.danger, row=1)
    async def close_round(
        self, interaction: discord.Interaction, button: discord.ui.Button
    ) -> None:
        round_row = await database.get_active_peer_round()
        if not round_row:
            await interaction.response.send_message("진행 중인 비밀평가가 없습니다.", ephemeral=True)
            return
        await database.close_peer_round(round_row["id"])
        await refresh_panel_after_close(self.bot, round_row)
        await interaction.response.send_message(
            embed=discord.Embed(
                title="🔒 비밀평가 종료",
                description=(
                    f"**{round_row['title']}** 라운드를 마감했습니다. "
                    "각 패널의 평가하기 버튼은 더 이상 작동하지 않습니다.\n"
                    "`/비밀평가 결과`로 최종 결과를 확인하세요."
                ),
                color=discord.Color.green(),
            ),
            ephemeral=True,
        )


async def open_admin_panel(interaction: discord.Interaction, bot: commands.Bot) -> None:
    """과제 대시보드의 [🔒 비밀평가 게시] 버튼 진입점(assignment cog에서 호출)."""
    round_row = await database.get_active_peer_round()
    status = f"진행 중 — **{round_row['title']}**" if round_row else "아직 시작되지 않음"
    await interaction.response.send_message(
        embed=discord.Embed(
            title="🔒 팀원 비밀평가 관리",
            description=(
                f"현재 상태: {status}\n\n"
                "• **팀 채널에 게시** — 각 팀 채널에 평가 패널을 올리고, 과제 대시보드에 완료율 진행판을 게시합니다.\n"
                "• **평가 미리보기** — 저장 없이 평가 화면을 테스트합니다(여기, 과제-대시보드 채널에서 가능).\n"
                "• **결과 보기 / 라운드 종료** — 관리자 전용."
            ),
            color=discord.Color.from_str(ACCENT),
        ),
        view=PeerEvalAdminView(bot),
        ephemeral=True,
    )


# ── 결과 집계 (관리자 전용) ─────────────────────────────────────────────────────

def aggregate_results(evals: list[dict]) -> dict[str, list[dict]]:
    """대상자별 평균 점수 집계. 반환: {팀: [ {name, count, avg1..4, overall, comments} ]}."""
    # target_id -> rows
    by_target: dict[str, list[dict]] = {}
    for e in evals:
        by_target.setdefault(e["target_id"], []).append(e)

    teams: dict[str, list[dict]] = {}
    for target_id, rows in by_target.items():
        n = len(rows)
        avgs = [sum(r[f"score{i + 1}"] for r in rows) / n for i in range(4)]
        overall = sum(avgs) / 4
        comments = [r["comment"].strip() for r in rows if r["comment"].strip()]
        team = rows[-1]["team"]
        teams.setdefault(team, []).append({
            "name": rows[-1]["target_name"],
            "count": n,
            "avgs": avgs,
            "overall": overall,
            "comments": comments,
        })
    for team in teams:
        teams[team].sort(key=lambda x: -x["overall"])
    return teams


def build_results_embeds(round_row: dict, results: dict[str, list[dict]]) -> list[discord.Embed]:
    header = discord.Embed(
        title=f"📊 비밀평가 결과 — {round_row['title']}",
        description=(
            "대상자별 **받은 평가 평균**입니다. 평가자 정보는 표시되지 않습니다(비밀평가).\n"
            f"지표: {' / '.join(f'{m}{n}' for m, n, _b, _a in INDICATORS)}"
        ),
        color=discord.Color.from_str(ACCENT),
    )
    embeds = [header]
    for team in TEAMS:
        rows = results.get(team)
        if not rows:
            continue
        lines = []
        for r in rows:
            scores = " ".join(f"{INDICATORS[i][0]}{r['avgs'][i]:.1f}" for i in range(4))
            lines.append(
                f"**{r['name']}** — 종합 **{r['overall']:.2f}** · {scores} · {r['count']}명 평가"
            )
        e = discord.Embed(
            title=f"👥 {team}",
            description="\n".join(lines)[:4000],
            color=discord.Color.from_str(ACCENT),
        )
        embeds.append(e)
    if len(embeds) == 1:
        header.add_field(name="​", value="아직 제출된 평가가 없습니다.", inline=False)
    return embeds[:10]


def build_results_excel(round_row: dict, results: dict[str, list[dict]]) -> tuple[io.BytesIO, str]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "비밀평가 결과"

    headers = [
        "팀", "대상자", "받은 평가 수",
        "① 과업완수", "② 협업소통", "③ 책임약속", "④ 팀기여", "종합 평균", "코멘트(취합)",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="8B5CF6")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for team in TEAMS:
        for r in results.get(team, []):
            ws.append([
                team, r["name"], r["count"],
                round(r["avgs"][0], 2), round(r["avgs"][1], 2),
                round(r["avgs"][2], 2), round(r["avgs"][3], 2),
                round(r["overall"], 2),
                "\n".join(f"- {c}" for c in r["comments"]),
            ])

    widths = [8, 16, 12, 12, 12, 12, 12, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        row[-1].alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_title = re.sub(r"[^\w가-힣-]", "_", round_row["title"])
    return buf, f"비밀평가결과_{safe_title}.xlsx"


async def _send_results(
    interaction: discord.Interaction, bot: commands.Bot, round_row: dict
) -> None:
    await interaction.response.defer(ephemeral=True)
    evals = await database.get_peer_evaluations(round_row["id"])
    results = aggregate_results(evals)
    embeds = build_results_embeds(round_row, results)
    file = None
    if evals:
        buf, fname = build_results_excel(round_row, results)
        file = discord.File(buf, filename=fname)
    await interaction.followup.send(
        embeds=embeds,
        file=file,
        ephemeral=True,
    )


# ── Cog ───────────────────────────────────────────────────────────────────────

class PeerEval(commands.Cog):
    def __init__(self, bot: commands.Bot) -> None:
        self.bot = bot

    async def cog_load(self) -> None:
        self.bot.add_view(PeerEvalPanelView(self.bot))

    group = app_commands.Group(name="비밀평가", description="팀원 비밀평가(동료평가) 관리 (관리자 전용)")

    @group.command(name="시작", description="비밀평가 라운드를 시작하고 과제 대시보드 채널에 패널을 게시합니다")
    @app_commands.describe(제목="라운드 제목 (예: 1차 팀원 상호평가). 생략 시 기본값")
    async def start(self, interaction: discord.Interaction, 제목: str | None = None) -> None:
        if not _is_admin(interaction):
            await interaction.response.send_message("관리자만 사용할 수 있습니다.", ephemeral=True)
            return
        guild = interaction.guild
        if not guild:
            await interaction.response.send_message("서버에서만 사용할 수 있습니다.", ephemeral=True)
            return
        ch = guild.get_channel(config.ASSIGNMENT_DASHBOARD_CHANNEL_ID)
        if not ch or not isinstance(ch, discord.TextChannel):
            await interaction.response.send_message(
                "과제 대시보드 채널을 찾을 수 없습니다.", ephemeral=True
            )
            return

        await interaction.response.defer(ephemeral=True)

        title = (제목 or "팀원 상호평가").strip()[:80]
        round_id = await database.create_peer_round(title)
        round_row = await database.get_peer_round(round_id)
        assert round_row is not None

        await post_progress_dashboard(self.bot, guild, round_row)

        roster = build_roster(guild)
        total = sum(len(v) for v in roster.values())
        roster_line = " · ".join(f"{t} {len(roster[t])}명" for t in TEAMS if roster[t])
        await interaction.followup.send(
            embed=discord.Embed(
                title="✅ 비밀평가 시작",
                description=(
                    f"**{title}** 라운드를 시작하고 {ch.mention} 채널에 패널을 게시했습니다.\n\n"
                    f"인식된 명단: 총 **{total}명** ({roster_line or '없음'})\n"
                    "닉네임에 팀 정보(예: `홍길동(3팀)`)가 없는 인원은 명단에서 빠집니다.\n"
                    "결과는 `/비밀평가 결과` 또는 패널의 **[📊 결과 보기]** 버튼으로 관리자만 확인합니다."
                ),
                color=discord.Color.green(),
            ),
            ephemeral=True,
        )

    @group.command(name="결과", description="비밀평가 집계 결과를 확인합니다 (관리자 전용, 엑셀 포함)")
    async def result(self, interaction: discord.Interaction) -> None:
        if not _is_admin(interaction):
            await interaction.response.send_message("관리자만 사용할 수 있습니다.", ephemeral=True)
            return
        round_row = await database.get_active_peer_round()
        if not round_row:
            await interaction.response.send_message(
                "진행 중이거나 마감된 비밀평가 라운드가 없습니다. `/비밀평가 시작`으로 시작하세요.",
                ephemeral=True,
            )
            return
        await _send_results(interaction, self.bot, round_row)

    @group.command(name="종료", description="현재 비밀평가 라운드를 마감합니다 (관리자 전용)")
    async def close(self, interaction: discord.Interaction) -> None:
        if not _is_admin(interaction):
            await interaction.response.send_message("관리자만 사용할 수 있습니다.", ephemeral=True)
            return
        round_row = await database.get_active_peer_round()
        if not round_row:
            await interaction.response.send_message(
                "진행 중인 비밀평가가 없습니다.", ephemeral=True
            )
            return
        await database.close_peer_round(round_row["id"])
        await refresh_panel_after_close(self.bot, round_row)
        await interaction.response.send_message(
            embed=discord.Embed(
                title="🔒 비밀평가 종료",
                description=(
                    f"**{round_row['title']}** 라운드를 마감했습니다. "
                    "패널의 평가하기 버튼은 더 이상 작동하지 않습니다.\n"
                    "`/비밀평가 결과`로 최종 결과를 확인하세요."
                ),
                color=discord.Color.green(),
            ),
            ephemeral=True,
        )


async def refresh_panel_after_close(bot: commands.Bot, round_row: dict) -> None:
    """마감 상태를 패널에 반영(활성 라운드가 사라지므로 전용 처리)."""
    panel = await database.get_assignment_panel(PANEL_TYPE)
    if not panel:
        return
    guild = bot.get_guild(config.GUILD_ID)
    if not guild:
        return
    ch = guild.get_channel(int(panel["channel_id"]))
    if not ch or not isinstance(ch, discord.TextChannel):
        return
    try:
        msg = await ch.fetch_message(int(panel["message_id"]))
        closed = dict(round_row)
        closed["is_active"] = 0
        evals = await database.get_peer_evaluations(round_row["id"])
        roster = build_roster(guild)
        embed = build_panel_embed(closed, roster, evals)
        await msg.edit(embed=embed, view=None)  # 버튼 제거
    except (discord.NotFound, discord.HTTPException) as e:
        log.warning("Peer-eval panel close-refresh failed: %s", e)


async def setup(bot: commands.Bot) -> None:
    await bot.add_cog(PeerEval(bot))
